import cleo

from exex import parse
from openpyxl import load_workbook

from exex_cli import convert
from exex_cli import formats


class ExtractCommand(cleo.Command):
    """
    Extract data from Excel document.

    extract
        {filename : Source file path}
        {--s|sheet=0 : Name of sheet}
        {--r|range=all : Range}
        {--f|format=text : text, table, json, csv}
        {--d|delimiter=, : Delimiter in output}
    """

    def handle(self):
        book = self.__get_book()
        sheet = self.__get_sheet(book)

        values_parsed = self.__parse_values_from_sheet(sheet)
        values_formatted = self.__format_values(values_parsed)
        self.__render_values(values_formatted)

    def __get_book(self):
        arg_filename = self.argument("filename")
        return load_workbook(arg_filename)

    def __get_sheet(self, book):
        arg_sheet = self.option("sheet")

        if arg_sheet == "0":
            arg_sheet = book.sheetnames[0]

        return book[arg_sheet]

    def __parse_values_from_sheet(self, sheet):
        arg_range = self.option("range")

        if arg_range == "all":
            values_raw = sheet.values
        else:
            values_raw = sheet[arg_range]

        return parse.values(values_raw)

    def __format_values(self, values):
        arg_format = self.option("format")

        if arg_format == formats.TEXT:
            return values
        elif arg_format == formats.CSV:
            return convert.to_csv(values, delimiter=self.option("delimiter"))
        elif arg_format == formats.TABLE:
            return convert.to_strings(values)
        elif arg_format == formats.JSON:
            return convert.to_json(values)
        else:
            return values

    def __render_values(self, values_formatted):
        arg_format = self.option("format")

        self.info("\n")

        if arg_format == formats.TABLE:
            self.render_table(headers=values_formatted[0], rows=values_formatted[1:])
        else:
            self.info(values_formatted)

        self.info("\n")

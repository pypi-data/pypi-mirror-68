import json
import os
import time
import string
import traceback

import openpyxl
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
import tablib
from openpyxl import Workbook, load_workbook, cell
from openpyxl.drawing.image import Image as ImageCell
from openpyxl.utils import get_column_letter
from swissarmykit.utils.fileutils import FileUtils
from swissarmykit.utils.stringutils import StringUtils
from swissarmykit.utils.timer import Timer

try: from definitions_prod import *
except Exception as e: pass # Surpass error. Note: Create definitions_prod.py

class ExcelUtils:

    def __init__(self, excel_path=None, data_only=False):
        if excel_path and os.path.exists(excel_path):
            self.file_name = excel_path
            if excel_path.endswith('.xls'):
                import xlrd
                self.wb = xlrd.open_workbook(excel_path)
            else:
                self.wb = openpyxl.load_workbook(excel_path, data_only=data_only)
            self.sheet_name = self.wb.get_sheet_names()

    def get_wb(self):
        return self.wb

    def get_active_sheet(self):
        return self.wb.active

    def save(self):
        new_file = self.file_name.replace('.xlsx', time.strftime(" - %Y-%m-%d %H-%M-%S", time.gmtime())+ '.xlsx')
        self.wb.save(new_file)
        print('\nINFO: Save to Excel %s' % new_file)

    @staticmethod
    def get_coordinate(column_index=1, row_index=1):
        '''
        Ex: C1
        :param column_index:  (3 -> 'C')
        :param row_index:     based 1
        :return: str
        '''
        return get_column_letter(column_index) + str(row_index)

    def get_sheet_by_name(self, name):
        return self.wb.get_sheet_by_name(name)

    def get_sheet_by_index(self, index):
        name = self.sheet_name[index]
        return self.wb.get_sheet_by_name(name)

    def print_sheet_info(self, active_ws):
        print('Max row: %d  - Max col: %d' % (active_ws.max_row, active_ws.max_column))

    def get_rows(self, min_row=1, max_row=None, max_col=None, sheet_index=0): # type: (int, int, int, int) -> [cell]
        '''This shit not test yet'''
        rows = []
        active_ws = self.get_sheet_by_index(sheet_index)
        self.print_sheet_info(active_ws=active_ws)
        for row in active_ws.iter_rows(min_row=min_row, max_row=max_row, max_col=max_col):
            rows.append(row)
        return rows

    def get_rows_of_col(self, col_start='A', col_end=None,  min_row=1, sheet_index=0):
        rows = []
        active_ws = self.get_sheet_by_index(sheet_index)
        self.print_sheet_info(active_ws=active_ws)

        if not col_end:
            col_end = col_start

        range_string = ('%s{}:%s{}' % (col_start, col_end)).format(min_row, active_ws.max_row)
        for row in active_ws.iter_rows(range_string):

            if col_start != col_end: # more than 2 columns
                cells = []
                for cell in row:
                    cells.append(cell)
                row.append(cells)
            else:
                for cell in row:
                    rows.append(cell.value)
        return rows

    def active_sheet(self, sheet_index):
        self.wb.active = sheet_index

    def write_at_coordination(self, cell_idx, value='example.com', style=None, hyperlink=None):
        self.wb.active[cell_idx].value = value

        if hyperlink:
            self.wb.active[cell_idx].hyperlink = hyperlink
            self.wb.active[cell_idx].style = 'Hyperlink'

        if style:
            self.wb.active[cell_idx].style = style

    def write_data_from(self, from_column='A', from_row=1, row_data=None):
        offset_col_ix = string.ascii_uppercase.index(from_column) + 1
        offset_row_ix = from_row

        if isinstance(row_data[0], dict):
            lst = []
            for row in row_data:
                lst.append([v for v in row.values()])
        else:
            lst = row_data


        for row_ix, row in enumerate(lst):
            for col_ix, col_val in enumerate(row):
                cell_idx = self.get_coordinate(offset_col_ix + col_ix, offset_row_ix + row_ix)
                self.write_at_coordination(cell_idx, value=col_val)

        print('Write file done')

    def write_into_sheet(self, active_ws,  rows):
        index = 0
        for row in rows:
            index += 1
            for cell in row:
                active_ws[cell.column + str(index)].value = cell.value
        print('Write file done')

    def validation(self, excel_path=None, json_path=None):
        if excel_path and '.xlsx' not in excel_path:
            raise Exception('Excel file must end with extension .xlsx')

        if json_path and '.json' not in json_path:
            raise Exception('Json file must end with extension .json')

    def json2excel(self, json_path=None):
        self.json2otherformat(json_path=json_path, format='xlsx')

    def __str__(self):
        from texttable import Texttable
        t = Texttable()
        lst = []
        for row in self.get_rows():
            lst.append([r.value for r in row])
        t.add_rows(lst)
        return t.draw()

    # @staticmethod
    # @deprecated
    # def json_obj_2_excel(json_obj, headers=None, json_path=None, format='xlsx', max_rows=65000): # xls
    #     '''  Use tablib.Dataset(*json_obj, headers=headers) '''
    #     # print(json_obj[0])
    #     if len(json_obj) > max_rows:
    #         print('ERROR: length is too big, ', len(json_obj))
    #         raise Exception('Can not export to Excel the big list. Please export into multiple Excel files.')
    #
    #     if isinstance(json_obj[0], dict):
    #         if not headers:
    #             headers = [header.replace('_', ' ') for header in json_obj[0].keys()]
    #         _lst = []
    #         for j in json_obj:
    #             _lst.append(list(j.values()))
    #         json_obj = _lst
    #
    #     if not headers:
    #         headers = ['' for i in range(0, len(json_obj[0]))]
    #
    #     if not json_path:
    #         json_path = appConfig.DIST_PATH + '/_excel/_json_obj_2_excel.' + format
    #     else:
    #         if not json_path.endswith(format):
    #             json_path += '.' + format
    #
    #     try:
    #         data = tablib.Dataset(*json_obj, headers=headers)
    #         FileUtils.write_binary_file(path=json_path, data=data.export(format))
    #         print('Output %s' % json_path)
    #         return True
    #     except Exception as e:
    #         if 'Permission denied:' in str(e):
    #             json_path +=  str(time.time()) + '.' + format
    #             data = tablib.Dataset(*json_obj, headers=headers)
    #             FileUtils.write_binary_file(path=json_path, data=data.export(format))
    #             print('The file is opening. New output would be %s' % json_path)
    #         else:
    #             print("ERROR: Somehow can not export Excel")
    #             print(e,'\n\n\n')
    #
    #             traceback.print_tb(e.__traceback__)
    #     return False

    @staticmethod
    def get_file_name(file_name=None, format='xlsx'):
        if not file_name:
            file_name = appConfig.DIST_PATH + '/_excel/_excel_data.' + format

        if '/' not in file_name:
            file_name = appConfig.EXCEL_PATH + '/' + file_name

        if '.' not in file_name.split('/')[-1]:
            file_name = file_name + '.' + format

        return file_name


    @staticmethod
    def json_to_openpyxl(json_obj, headers_order=None, headers=None, file_name=None, format='xlsx', append=False, write_one_by_one=False,
                         remove_invalid_utf8=False, callback_get_valid_utf8=None, style_header=True): # xls
        ''' Use  from openpyxl import Workbook'''
        _total_rows = len(json_obj)
        print('INFO: Total rows: ', _total_rows)
        timer = Timer.instance(_total_rows)
        is_row_obj = isinstance(json_obj[0], dict)

        file_name = ExcelUtils.get_file_name(file_name)

        if not append:
            # Step 1: Header empty or default
            if headers_order:
                headers = headers_order
            elif not headers:
                headers = [header for header in json_obj[0].keys()] if is_row_obj else [''] * len(json_obj[0])

            # Step 2: Reformat list result
            if is_row_obj:
                _lst = []
                for j in json_obj:
                    _lst.append([j.get(h, '') for h in headers_order] if headers_order else list(j.values()))
                json_obj = _lst

        # Extract to.
        try:
            from openpyxl import Workbook
            if append:
                wb = load_workbook(file_name)
            else:
                wb = Workbook()
            ws = wb.active
            ws.append(list(headers))

            if style_header:
                bold_font = Font(color='FFFFFF', bold=True)
                blue = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
                for cell in ws['1:1']:
                    cell.font  = bold_font
                    cell.fill = blue

            if write_one_by_one:
                row_idx = 1
                for row in json_obj:
                    row_idx += 1
                    timer.check()
                    for idx, cell in enumerate(row):
                        cell_idx = ExcelUtils.get_coordinate(idx + 1, row_idx)
                        if isinstance(cell, dict):

                            if cell.get('hyperlink'):
                                ws[cell_idx].value = StringUtils.get_valid_utf_8(cell.get('value')) if remove_invalid_utf8 else cell.get('value')
                                ws[cell_idx].hyperlink = cell.get('hyperlink')
                                ws[cell_idx].style = 'Hyperlink'

                            elif cell.get('image'):
                                if cell.get('height'):
                                    ws.row_dimensions[row_idx].height = cell.get('height')

                                img = ImageCell(cell.get('image'))
                                img.anchor = cell_idx
                                ws.add_image(img)
                        else:
                            ws[cell_idx].value = cell if cell else ''
            else:
                count = 1
                for row in json_obj:
                    try:
                        timer.check()
                        if remove_invalid_utf8:
                            row = [StringUtils.get_valid_utf_8(r) for r in row]

                        if callback_get_valid_utf8:
                            row = [callback_get_valid_utf8(r) for r in row]

                        ws.append(row)
                        count += 1
                    except Exception as e:
                        try:
                            row = [StringUtils.get_valid_utf_8(r) for r in row]
                            print('INFO: retry append at ID: ' + str(count))
                            ws.append(row)
                            count += 1
                        except Exception as e1:
                            print('ERROR: even retry', e1, count, row)

            wb.save(file_name)
            print('\nINFO: Output to Excel %s\n' % file_name)
            return file_name
        except Exception as e:
            if 'Permission denied:' in str(e):
                file_name += str(time.time()) + '.' + format
                wb.save(file_name)
                # data = tablib.Dataset(*json_obj, headers=headers)
                # FileUtils.write_binary_file(path=file_name, data=data.export(format))
                print('The file is opening. New output would be %s' % file_name)
            else:
                print("ERROR: Somehow can not export Excel")
                print(e, '\n\n\n')

                traceback.print_tb(e.__traceback__)
        return ''

    def json2otherformat(self, json_path=None, format='xlsx'):
        ''' https://pypi.org/project/json2xlsx/
        '''
        self.validation(json_path=json_path)
        data = json.loads(FileUtils.read_file(json_path))
        if type(data) == list:
            headers = list(data[0].keys())
            data = [list(v.values()) for v in data]
            data = tablib.Dataset(*data, headers=headers)

            output = json_path.replace('.json', '_out.' + format)
            if format == 'xlsx':
                FileUtils.write_binary_file(path=output, data=data.export(format))
            else:
                FileUtils.write_file(path=output, data=data.export(format))

            print('Output %s' % output)
            return True

        return False

    def set_cell(self, col_input_idx, value):
        self.wb.active[col_input_idx] = value

    def save_change(self):
        output = self.file_name.replace('.xlsx', '_output.xlsx')
        self.wb.save(output)
        print('Save change success: %s' % output)


    def fill_column(self, col_input_idx='A', col_ouput_idx='B', callback=None, excel_path=None, start_at=2, stop_at=None):
        if not callback or not excel_path:
            raise Exception('Please input callback func and excel_path.')

        self.validation(excel_path=excel_path)

        book = openpyxl.load_workbook(excel_path)
        sheet = book.active

        if not stop_at:
            stop_at = sheet.max_row

        for i in range(start_at, stop_at + 1):
            input_i = col_input_idx + str(i)
            input = sheet[input_i].value

            output_i = col_ouput_idx + str(i)
            output = callback(input)
            sheet[output_i] = output
            print('[%s] = %s   =>  [%s] = %s' % (input_i, input, output_i, output))

        book.save(excel_path.replace('.xlsx', '_output.xlsx'))


    def classify_into_2_groups(self, classify_callback=None, classify_agaist=None, rows=None):
        '''
        Example:

        excel = ExcelUtils(appConfig.dist_path('manipulate_2.xlsx'))

        all_email = excel.get_rows_of_col(col_start='E', min_row=2, sheet_index=3)
        set_email = set(all_email)
        def compare_equal(row, classify_agaist):
            return row[2].value in classify_agaist

        mortage = excel.get_rows(1, sheet_index=0)
        texax = excel.get_rows(2, sheet_index=1)

        group_1, group_5 = excel.classify_into_2_groups(compare_equal, set_email, mortage)
        group_2, group_6 = excel.classify_into_2_groups(compare_equal, set_email, texax)
        rows = group_1 + group_2

        excel.write_into_sheet(sheet3, rows)
        excel.write_into_sheet(sheet5, group_5)
        excel.write_into_sheet(sheet6, group_6)
        excel.save()
        :return:
        '''
        group_1 = []
        group_2 = []
        for row in rows:
            if classify_callback(row, classify_agaist):
                group_1.append(row)
            else:
                group_2.append(row)

        return (group_1, group_2)


if __name__ == '__main__':
    # excel = ExcelUtils(appConfig.dist_path('manipulate_2.xlsx'))
    #
    # all_email = excel.get_rows_of_col(col_start='E', min_row=2, sheet_index=3)
    # set_email = set(all_email)
    #
    #
    # sheet3 = excel.get_sheet_by_index(2)
    # sheet5 = excel.get_sheet_by_index(4)
    # sheet6 = excel.get_sheet_by_index(5)
    #
    # def compare_equal(a, b):
    #     return a[2].value in b
    #
    # mortage = excel.get_rows(1, sheet_index=0)
    # texax = excel.get_rows(2, sheet_index=1)
    #
    # group_1, group_5 = excel.classify_into_2_groups(compare_equal, set_email, mortage)
    # group_2, group_6 = excel.classify_into_2_groups(compare_equal, set_email, texax)
    # rows = group_1 + group_2
    #
    # excel.write_into_sheet(sheet3, rows)
    # excel.write_into_sheet(sheet5, group_5)
    # excel.write_into_sheet(sheet6, group_6)
    # excel.save()

    # path = appConfig.DIST_PATH + '/input.json'
    # excel.json2otherformat(path, format='yaml' )
    pass
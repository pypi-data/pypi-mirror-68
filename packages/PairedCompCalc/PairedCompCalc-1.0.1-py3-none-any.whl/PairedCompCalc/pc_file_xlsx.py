"""Read paired-comparison data stored in an Excel 2007- (xlsx) workbook file.

Each Excel work-book file, and each work-sheet in the file,
may include paired-comparison results for
one or more subject(s), tested in one or more test-factor combinations.

Each row in the worksheet must include enough data to specify
one stimulus pair and the corresponding (possibly graded) judgment response,
to be represented by ONE PairedCompItem instance.

The perceived difference between the two sounds in a presented pair (stim0, stim1)
may be encoded in three different ways:
*1: a single signed integer in (-M, ..., +M) showing the signed difference,
    where a positive value means that stim1 was the "better" (preferred) sound,
*2: a combination of (choice, response_label), where
    choice is a string == either stim0 or stim1,
    or a string in the parameter no_choice, indicating no perceived difference.
    The response_label is a string identifying the magnitude of the difference.
    The response_label must be an element in the list PairedCompFrame.difference_grades.
*3: only the string choice, if the judgment is binary, i.e., with no difference grading.

A judgment of "no difference" can be represented EITHER by
choice == an element in the list no_choice, OR
response_label == PairedCompFrame.difference_grades[0], if forced_choice == False.


*** Usage Example:

pcf=PairedCompFrame(objects=['HA-1', 'HA-2', 'HA-3'],
            attributes=['Preference'],
            difference_grades=['equal', 'slightly better', 'better', 'much better'],
            test_factors=dict(Sound=['speech', 'music'], SNR=['low', 'high']),
            forced_choice=False
            )

pc_file = PairedCompFile(file_path,
        pcf=pcf,
        sheets=[f'Subject{i} for i in range(10)],
        subject='sheet',
        top_row=4,  # first row with paired-comparison data
        attribute='A',
        pair=('B', 'C'),
        difference='D',
        choice='E',
        Sound='G'  # category for Test Factor Sound, as defined in pcf.test_factors
        )

for pc_item in pc_file:
    process the pc_item

The parameter pcf defines experimental layout and
possible categories within each given test factor to be analyzed.

The parameter sheets is a list of workbook sheets to be searched for results.

Other PairedCompFile properties define locations for
subject, attribute, stimulus-pair labels, difference magnitude,
and test factors.

The location of subject ID, attribute, and test factors can be a column address
OR
'sheet', indicating that the sheet name is to be used for all included items.

In this example, subject='sheet' indicates that the sheet name is interpreted as subject id.
pair=('B', 'C') defines columns containing stimulus labels.
This parameter may be omitted if pcf.objects includes exactly TWO elements,
and then it is assumed that every pair == pcf.objects.

attribute='A' defines the column defining the attribute label.
This parameter may be omitted if pcf.attributes includes exactly ONE element,
as in this example.

Thus, in this example, each row x, for x >= 4, specifies
the pair (stim0, stim1) in cells Bx and Cx,
the choice in cell Ex,
the corresponding response label in cell Dx,
and cell Gx specifies a category for test-factor 'Sound'.

The file path string will be examined to find
one of the allowed categories for the remaining test-factor 'SNR',
which does not have a column address.


*** Version History:

2018-10-01, first functional version
"""

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.utils import column_index_from_string
import logging

from PairedCompCalc import pc_base

logger = logging.getLogger(__name__)
# logger.setLevel(logging.DEBUG)  # TEST


class FileFormatError(pc_base.FileReadError):
    """Format error causing non-usable data"""


class ParameterError(pc_base.FileReadError):
    """Error in calling parameters causing non-usable data"""


class ArgumentError(RuntimeError):
    """Error in calling parameters causing non-usable data"""


class PairedCompFile(pc_base.PairedCompFile):
    """Interface to Excel xlsx workbook file storing paired-comparison data.
    Each sheet in the workbook may include data for
    one or more subjects, attributes, and test conditions.
    Data elements MUST be stored at the same location in all sheets.
    """
    def __init__(self, file_path, pcf,
                 subject=None,
                 attribute=None,
                 pair=None,
                 difference=None,
                 choice=None,
                 no_choice=None,
                 top_row=2,
                 sheets=None,
                 **test_factors):
        """File interface to data stored in an excel file,
        decoded as a sequence of PairedCompItem instances.
        :param file_path: Path to file for reading
        :param pcf: single PairedCompFrame instance
        :param subject: 'sheet' or column for subject identification label
        :param attribute: (optional) 'sheet' or column for perceptual attribute labels
            may be None if len(pcf.attributes) == 1
        :param pair: (optional) tuple with two column addresses defining (stim0, stim1)
            may be None if there are only two objects in total for comparison
        :param difference: column with perceived difference, coded as EITHER
            a string label for the (unsigned) difference magnitude, OR
            an integer indicating the ranked difference of stim1 rel. stim0,
            i.e., with positive value indicating that stim1 was better than stim0
        :param choice: (optional) column with label of preferred pair element
            if None, choice is specified by sign of difference
        :param no_choice: list of possible cell values in the 'choice' column,
            indicating no audible difference or no preference
        :param top_row: integer address of first row containing PairedCompItem data
        :param sheets: list of sheet names to be searched for data.
        :param test_factors: (optional) dict with elements (tf, location), where
            tf is a string identifying one test factor,
            location is a string with a column (top) address like 'D',
            OR 'sheet', indicating that the sheet name is the test-factor category.
        """
        super().__init__(file_path, pcf)
        self.subject = _check_column_or_sheet(subject)
        self.attribute = _check_column_or_sheet(attribute)
        self.pair = pair  # ****************************** Check!
        self.choice = _check_column(choice)
        self.no_choice = no_choice  # *******************
        self.difference = _check_column(difference)
        self.top_row = top_row
        self.sheets = sheets
        self.test_factors = _check_test_cond(test_factors)

    def __iter__(self):
        """Generator yielding data from an excel file,
        in dicts compatible with class StimRespRecord.
        :return: generator yielding PairedCompItem instance
        """
        try:
            wb = load_workbook(str(self.file_path), read_only=True)
        except InvalidFileException:
            raise FileFormatError(f'Cannot load workbook from file {self.file_path.stem}')
        if self.sheets is None:
            sheets = wb.sheetnames
        else:
            sheets = set(self.sheets) & set(wb.sheetnames)
        if len(sheets) == 0:
            raise FileFormatError('No accepted sheets found in {file_path.stem}')
        path_test_cond = self.path_test_cond()
        for sheet_name in sheets:
            ws = wb[sheet_name]
            rows = ws.rows
            for _ in range(self.top_row - 1):
                row = rows.__next__()
                logger.debug(f'skipping row {row[0].row}')
            for row in rows:
                p = self._get_pair(ws, row)
                r = pc_base.PairedCompItem(subject=self._get_subject(ws, row),
                                           attribute=self._get_attribute(ws, row),
                                           pair=p,
                                           response=self._get_response(ws, row, p),
                                           test_cond=self._get_test_cond(ws, row, path_test_cond))
                if not any(r_i is None for r_i in r.__dict__.values()):
                    # we have a valid item
                    yield r
                else:
                    logger.debug(f'not using row {row[0]}...')

    def _get_subject(self, ws, row):
        """
        :param ws: a worksheet
        :param row: tuple of openpyxl Cell instances
        :return: subject code, integer or string
        """
        return _get_value(ws, self.subject, row)

    def _get_attribute(self, ws, row):
        """
        :param ws: a worksheet
        :param row: tuple of openpyxl Cell instances
        :return: string label
        """
        if self.attribute is None:
            if len(self.pcf.attributes) == 1:
                return self.pcf.attributes[0]
            else:
                raise FileFormatError(f'Undefined attribute in file {self.file_path.stem}')
        else:
            return _get_value(ws, self.attribute, row)

    def _get_pair(self, ws, row):
        """
        :param ws: a worksheet
        :param row: tuple of openpyxl Cell instances
        :return: tuple of strings (stim0, stim1)
        """
        if self.pair is None:
            if len(self.pcf.objects) == 2:
                return tuple(self.pcf.objects)
            else:
                raise FileFormatError(f'Undefined pair in file {self.file_path.stem}')
        stim0 = _get_value(ws, self.pair[0], row)
        stim1 = _get_value(ws, self.pair[1], row)
        if stim0 is None and stim1 is None:
            return None
        else:
            return stim0, stim1

    def _get_response(self, ws, row, pair):
        """Encode response as an integer in (-M, ..., +M)
        Responses may be code either as (choice, difference_magnitude)
        OR simply as difference_signed with sign indicating the choice
        :param ws: a worksheet
        :param row: tuple of openpyxl Cell instances
        :param pair: tuple (stim0, stim1), either default or from this row
        :return: integer in (-M,...,M) indicating ranked difference
        """
        if self.choice is None:
            return _get_value(ws, self.difference, row)
        else:
            choice = _get_value(ws, self.choice, row)
        if choice is None:
            return None
        if self.no_choice is not None and choice in self.no_choice:
            return 0
        if choice == pair[0]:
            response_sign = -1
        elif choice == pair[1]:
            response_sign = 1
        else:
            return None
        if self.difference is None:
            return response_sign
        else:
            response_label = _get_value(ws, self.difference, row)
            try:
                response_magn = self.pcf.difference_grades.index(response_label)
                return response_sign * response_magn
            except ValueError:
                return None

    def _get_test_cond(self, ws, row, path_tc):
        """Get tuple of test conditions
        :param ws: a worksheet
        :param row: tuple of openpyxl Cell instances
        :param path_tc: default dict with (tf, tf_category) from file path
        :return: dict with elements (tf, tf_category)
        """
        tc = path_tc.copy()
        for (tf, col) in self.test_factors.items():
            tc[tf] = _get_value(ws, col, row)
        return tc


# --------------------------------------------------- help sub-functions

def _check_column(col):
    """Check that a parameter address is acceptable
    """
    if col is None or type(col) is str and col.isupper():
        return col
    # otherwise:
    raise ArgumentError(f'Column address {col} must be string of uppercase letters')


def _check_column_or_sheet(col):
    """Check that a parameter address is either 'sheet' or a column string address
    """
    if type(col) is str and col == 'sheet':
        return col
    else:
        return _check_column(col)


def _check_test_cond(tf_address):
    for (tf, col) in tf_address.items():
        _check_column_or_sheet(col)
    return tf_address


def _get_value(ws, col, row):
    """Get contents in ONE cell or in sheet title
    :param ws: a worksheet
    :param col: one column address or 'sheet'
    :param row: integer address of row
    :return: cell contents
    """
    if col == 'sheet':
        return ws.title
    # c = ws[col + str(row)]
    cell = row[column_index_from_string(col) - 1]
    return cell.value



import os
import openpyxl
import glob
"""sorry for bad english"""
"""i made this when im itern in qtdatagroup, recieve a task to get data from some clsx file"""
"""this module has some functions which give you header and data of some xlsx file, with bonus
    3 column: filename, field, city, with filename = filename, field = purchase, city = hanoi or hcm"""
#get sheet_obj
def get_list_sheet_obj(path):
    """this function return a list of worksheet object of a workbook,
        which has path as an argument"""
    result=[]
    wb_obj = openpyxl.load_workbook(path)
    for _sheet_name in wb_obj.sheetnames:
        sheet_obj = wb_obj[_sheet_name]
        result.append(sheet_obj)
    return result
#lay danh sach tham so header
def get_header_excel_files(path):
    """this function take arrgumant called path which is 
        path of file, then return a list of header"""
    wb2 = openpyxl.load_workbook(path)
    sheet_obj_base = wb2.active
    result = []
    for column in range(1, sheet_obj_base.max_column+1):
        cell_obj = sheet_obj_base.cell(1, column)      
        result.append(cell_obj.value)
    bonus_header = ["filename", "field", "city"]
    result.extend(bonus_header)
    return result

#get data
def get_data_of_workbook(worksheet_obj, _row, path, _field, _city):
    """this function take argument is worksheet_obj, which is connect to workbook
        , full path of workbook
        ,_row(dont input 1, 1 is header), 
        _filed = Purchase,
        _city = Hanoi or HCM base on your task
        return list of data in a row"""
    result = []
    for column in range(1, worksheet_obj.max_column+1):
        cell_obj = worksheet_obj.cell(_row, column)      
        result.append(cell_obj.value)
    bonus_data = [path, _field, _city]
    result.extend(bonus_data)
    return result
#get full path of files
def dirwalk(dir, bag, wildcards):
    """this function return a list of full path of files or folder in folder dir with condition wildcard
        bag: result list"""
    bag.extend(glob.glob(os.path.join(dir, wildcards)))
    for f in os.listdir(dir):
        fullpath = os.path.join(dir, f)
        if os.path.isdir(fullpath) and not os.path.islink(fullpath):
            dirwalk(fullpath, bag, wildcards)
#check if a cell is empty
def check_empty_value(sheet_obj, row, column):
    cell_obj = sheet_obj.cell(row, column)
    if cell_obj.value is None:
        return True
    return False
#get real hightest row
def get_highest_row(sheet_obj):
    nums_row = 0
    for row in range(1, sheet_obj.max_row+1):
        if check_empty_value(sheet_obj, row, 5):
            return nums_row
        nums_row = nums_row + 1
    return nums_row
def complete_project(source_path, result_path, field, city):
    """this is function to complete this task, which
        source_path is source of folder
        result_path is path of result
        field, in this task, it is "Purchase"
        city, is Ha Noi or Ho Chi Minh"""
    wb_result = openpyxl.load_workbook(result_path)
    #get list of full path of files in folder
    list_source_path=[]
    dirwalk(source_path, list_source_path, "*.xlsx")
    #get header
    list_header=[]
    list_header=get_header_excel_files(list_source_path[0])
    #add header to result file
    ws_result = wb_result.active
    ws_result.append(list_header)
    for _file in list_source_path:
        list_ws_source=[] 
        list_ws_source = get_list_sheet_obj(_file)
        for _ws_source in list_ws_source:
            #get real max rows
            real_max_row = get_highest_row(_ws_source)
            for _row in range(2, real_max_row):
                _result = []
                _result = get_data_of_workbook(_ws_source, _row, _file, field, city)
                ws_result.append(_result)
    print(ws_result.max_row)        
        
    wb_result.save(result_path)
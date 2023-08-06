# -*- coding: UTF-8 -*-
#! python3

"""
    Get metadatas from Isogeo and store it into a Excel worksheet.

"""

# ##############################################################################
# ########## Libraries #############
# ##################################

# Standard library
import logging
from collections.abc import KeysView
from pathlib import Path
from urllib.parse import urlparse

# 3rd party library
from isogeo_pysdk import IsogeoUtils, Metadata, Share
from openpyxl import Workbook
from openpyxl.styles import Alignment, NamedStyle
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# custom submodules
from isogeotoxlsx.i18n import I18N_EN, I18N_FR
from isogeotoxlsx.matrix import (
    ColumnPattern,
    ATTRIBUTES_COLUMNS,
    RASTER_COLUMNS,
    RESOURCE_COLUMNS,
    SERVICE_COLUMNS,
    VECTOR_COLUMNS,
)
from isogeotoxlsx.utils import Formatter, Stats

# ##############################################################################
# ############ Globals ############
# #################################

logger = logging.getLogger("isogeotoxlsx")
utils = IsogeoUtils()

# ##############################################################################
# ########## Classes ###############
# ##################################


class Isogeo2xlsx(Workbook):
    """Used to store Isogeo API results into an Excel worksheet (.xlsx)

    :param str lang: selected language for output
    :param str url_base_edit: base url to format edit links (basically app.isogeo.com)
    :param str url_base_view: base url to format view links (basically open.isogeo.com)
    """

    def __init__(
        self,
        lang: str = "FR",
        url_base_edit: str = "",
        url_base_view: str = "",
        # additional
        **kwargs,
    ):
        """Instanciating the output workbook.

        :param str lang: selected language for output
        :param str url_base_edit: base url to format edit links (basically app.isogeo.com)
        :param str url_base_view: base url to format view links (basically open.isogeo.com)
        """
        super(Isogeo2xlsx, self).__init__()
        # super(Isogeo2xlsx, self).__init__(write_only=True)

        self.stats = Stats()

        # URLS
        utils.app_url = url_base_edit  # APP
        utils.oc_url = url_base_view  # OpenCatalog url

        # styles
        s_date = NamedStyle(name="date")
        s_wrap = NamedStyle(name="wrap")
        s_wrap.alignment = Alignment(wrap_text=True)
        self.add_named_style(s_date)
        self.add_named_style(s_wrap)

        # deleting the default worksheet
        ws = self.active
        self.remove(ws)

        # LOCALE
        if lang.lower() == "fr":
            s_date.number_format = "dd/mm/yyyy"
            self.dates_fmt = "DD/MM/YYYY"
            self.locale_fmt = "fr_FR"
            self.tr = I18N_FR
        else:
            s_date.number_format = "yyyy/mm/dd"
            self.dates_fmt = "YYYY/MM/DD"
            self.locale_fmt = "uk_UK"
            self.tr = I18N_EN

        # FORMATTER
        self.fmt = Formatter(output_type="Excel")

    # ------------ Setting workbook ---------------------

    def set_worksheets(
        self,
        auto: KeysView = None,
        vector: bool = 1,
        raster: bool = 1,
        service: bool = 1,
        resource: bool = 1,
        dashboard: bool = 0,
        attributes: bool = 0,
        fillfull: bool = 0,
        inspire: bool = 0,
    ):
        """Adds new sheets depending on present metadata types in isogeo API
        search tags.

        :param list auto: typically auto=search_results.get('tags').keys()
        :param bool vector: add vector sheet
        :param bool raster: add raster sheet
        :param bool service: add service sheet
        :param bool resource: add resource sheet
        :param bool dashboard: add dashboard sheet
        :param bool attributes: add attributes sheet - only if vector is True too
        :param bool fillfull: add fillfull sheet
        :param bool inspire: add inspire sheet
        """
        if isinstance(auto, KeysView):
            logger.info("Automatic sheets creation based on tags")
            if "type:vector-dataset" in auto:
                vector = 1
                self.columns_vector = {
                    k: ColumnPattern._make(v) for k, v in VECTOR_COLUMNS.items()
                }
            else:
                vector = 0
            if "type:raster-dataset" in auto:
                raster = 1
                self.columns_raster = {
                    k: ColumnPattern._make(v) for k, v in RASTER_COLUMNS.items()
                }
            else:
                raster = 0
                pass
            if "type:resource" in auto:
                resource = 1
                self.columns_resource = {
                    k: ColumnPattern._make(v) for k, v in RESOURCE_COLUMNS.items()
                }
            else:
                resource = 0
                pass
            if "type:service" in auto:
                service = 1
                self.columns_service = {
                    k: ColumnPattern._make(v) for k, v in SERVICE_COLUMNS.items()
                }
            else:
                service = 0
                pass
        else:
            raise TypeError(
                "'auto' must be a KeysView (dict.keys()),"
                " from Isogeo search request, not {}".format(type(auto))
            )

        # SHEETS & HEADERS
        if dashboard:
            self.ws_d = self.create_sheet(title=self.tr.get("dashboard"))
            # initialize line counter
            self.idx_d = 1
            # log
            logger.info("Dashboard sheet added")
        else:
            pass
        if fillfull:
            self.ws_f = self.create_sheet(title="Progression catalogage")
            # headers
            # self.ws_f.append([i for i in self.cols_v])
            # styling
            # for i in self.cols_v:
            #     self.ws_v.cell(row=1,
            #                    column=self.cols_v.index(i) + 1).style = "Headline 2"
            # initialize line counter
            self.idx_f = 1
            # log
            logger.info("Fillfull sheet added")
        else:
            pass
        if inspire:
            self.ws_i = self.create_sheet(title="Directive INSPIRE")
            # headers
            # self.ws_f.append([i for i in self.cols_v])
            # styling
            # for i in self.cols_v:
            #     self.ws_v.cell(row=1,
            #                    column=self.cols_v.index(i) + 1).style = "Headline 2"
            # initialize line counter
            self.idx_i = 1
            # log
            logger.info("INSPIRE sheet added")
        else:
            pass
        if vector:
            self.ws_v = self.create_sheet(title=self.tr.get("vector"))
            # headers
            self.headers_writer(ws=self.ws_v, columns=self.columns_vector)
            # initialize line count
            self.idx_v = 1
            # log
            logger.info("Vectors sheet added")
            # feature attributes analisis
            if attributes:
                # columns matching table
                self.columns_attributes = {
                    k: ColumnPattern._make(v) for k, v in ATTRIBUTES_COLUMNS.items()
                }
                # create worksheet
                self.ws_fa = self.create_sheet(title=self.tr.get("attributes"))
                # set index
                self.idx_fa = 1
                self.fa_all = []
                # headers
                self.headers_writer(ws=self.ws_fa, columns=self.columns_attributes)
                # log
                logger.info("Feature attributes sheet added")
            else:
                pass
        else:
            pass

        if raster:
            self.ws_r = self.create_sheet(title=self.tr.get("raster"))
            # headers
            self.headers_writer(ws=self.ws_r, columns=self.columns_raster)
            # initialize line counter
            self.idx_r = 1
            # log
            logger.info("Rasters sheet added")
        else:
            pass

        if service:
            self.ws_s = self.create_sheet(title=self.tr.get("service"))
            # headers
            self.headers_writer(ws=self.ws_s, columns=self.columns_service)
            # initialize line counter
            self.idx_s = 1
            # log
            logger.info("Services sheet added")
        else:
            pass

        if resource:
            self.ws_rz = self.create_sheet(title=self.tr.get("resource"))
            # headers
            self.headers_writer(ws=self.ws_rz, columns=self.columns_resource)
            # initialize line counter
            self.idx_rz = 1
            # log
            logger.info("Resources sheet added")
        else:
            pass

    # ------------ Writing metadata ---------------------
    def store_metadatas(self, metadata: Metadata, share: Share = None):
        """Write metadata into the worksheet.

        :param Metadata metadata: metadata object to write
        :param Share share: share to use to build the OpenCatalog URL
        """
        # check input
        if not isinstance(metadata, Metadata):
            raise TypeError("Export expects a Metadata object.")
        # generic export
        self.share = share
        # store depending on metadata type
        if metadata.type == "vectorDataset":
            self.idx_v += 1
            self.store_md_generic(metadata, self.ws_v, self.idx_v)
            self.stats.md_types_repartition["vector"] += 1
            # style it
            self.column_width(self.ws_v, self.columns_vector)
            self.row_height(self.ws_v)
            self.styling_cells(self.ws_v, self.columns_vector)
        elif metadata.type == "rasterDataset":
            self.idx_r += 1
            self.store_md_generic(metadata, self.ws_r, self.idx_r)
            self.stats.md_types_repartition["raster"] += 1
            # style it
            self.column_width(self.ws_r, self.columns_raster)
            self.row_height(self.ws_r)
            self.styling_cells(self.ws_r, self.columns_raster)
        elif metadata.type == "service":
            self.idx_s += 1
            self.store_md_generic(metadata, self.ws_s, self.idx_s)
            self.stats.md_types_repartition["service"] += 1
            # style it
            self.column_width(self.ws_s, self.columns_service)
            self.row_height(self.ws_s)
            self.styling_cells(self.ws_s, self.columns_service)
        elif metadata.type == "resource":
            self.idx_rz += 1
            self.store_md_generic(metadata, self.ws_rz, self.idx_rz)
            self.stats.md_types_repartition["resource"] += 1
            # style it
            self.column_width(self.ws_rz, self.columns_resource)
            self.row_height(self.ws_rz)
            self.styling_cells(self.ws_rz, self.columns_resource)
        else:
            logger.error(
                "Type of metadata is not recognized/handled: {}".format(metadata.type)
            )

    def store_md_generic(self, md: Metadata, ws: Worksheet, idx: int):
        """Exports generic metadata attributes into Excel worksheet with some dynamic
        adaptations based on metadata type.

        :param Metadata md: metadata object to export
        :param Worksheet ws: Excel worksheet to store the exported info
        :param int idx: row index in the worksheet
        """
        # pick columns referential table depending on metadata type
        if md.type == "rasterDataset":
            col = self.columns_raster
        elif md.type == "resource":
            col = self.columns_resource
        elif md.type == "service":
            col = self.columns_service
        elif md.type == "vectorDataset":
            col = self.columns_vector
        else:
            raise TypeError("Unknown metadata type: {}".format(md.type))

        logger.debug(
            "Start storing metadata {} ({}) using the matching reference columns for type of {} ...".format(
                md.title_or_name(slugged=1), md._id, md.type
            )
        )

        # -- IDENTIFICATION ------------------------------------------------------------
        if md.title:
            ws["{}{}".format(col.get("title").letter, idx)] = md.title
        if md.name:
            ws["{}{}".format(col.get("name").letter, idx)] = md.name
        if md.abstract:
            ws["{}{}".format(col.get("abstract").letter, idx)] = md.abstract

        # path to source
        try:
            src_path = Path(str(md.path))
        except OSError as e:
            logger.debug(
                "Metadata.path value is not a valid system path. Maybe an URL? Original error: {}".format(
                    e
                )
            )
            urlparse(md.path).scheme != ""

        if isinstance(md.path, Path) and md.type != "service":
            if src_path.is_file():
                link_path = r'=HYPERLINK("{0}","{1}")'.format(
                    src_path.parent, src_path.resolve()
                )
                ws["{}{}".format(col.get("path").letter, idx)] = link_path
                logger.debug("Path reachable: {}".format(src_path))
            else:
                ws["{}{}".format(col.get("path").letter, idx)] = str(src_path.resolve())
                logger.debug("Path unreachable: {}".format(str(src_path)))
        elif md.path and md.type == "service":
            link_path = r'=HYPERLINK("{0}","{1}")'.format(md.path, md.path)
            ws["{}{}".format(col.get("path").letter, idx)] = link_path
        elif md.path:
            ws["{}{}".format(col.get("path").letter, idx)] = md.path
            logger.debug("Path not recognized: {}".format(str(src_path)))
        else:
            pass

        # -- TAGS ----------------------------------------------------------------------
        keywords = []
        inspire = []
        if md.keywords:
            for k in md.keywords:
                if k.get("_tag").startswith("keyword:is"):
                    keywords.append(k.get("text"))
                elif k.get("_tag").startswith("keyword:in"):
                    inspire.append(k.get("text"))
                else:
                    logger.info("Unknown keyword type: " + k.get("_tag"))
                    continue
            if keywords:
                ws["{}{}".format(col.get("keywords").letter, idx)] = " ;\n".join(
                    sorted(keywords)
                )
            if inspire:
                ws["{}{}".format(col.get("inspireThemes").letter, idx)] = " ;\n".join(
                    sorted(inspire)
                )
        else:
            self.stats.md_empty_fields[md._id].append("keywords")
            logger.info("Vector dataset without any keyword or INSPIRE theme")

        # INSPIRE conformity
        if col.get("inspireConformance").letter is not None:
            ws["{}{}".format(col.get("inspireConformance").letter, idx)] = (
                "conformity:inspire" in md.tags
            )

        # owner
        ws["{}{}".format(col.get("_creator").letter, idx)] = next(
            v for k, v in md.tags.items() if "owner:" in k
        )

        # -- HISTORY -------------------------------------------------------------------
        if md.collectionContext:
            ws[
                "{}{}".format(col.get("collectionContext").letter, idx)
            ] = md.collectionContext
        if md.collectionMethod:
            ws[
                "{}{}".format(col.get("collectionMethod").letter, idx)
            ] = md.collectionMethod

        # validity
        if md.validFrom:
            ws["{}{}".format(col.get("validFrom").letter, idx)] = utils.hlpr_datetimes(
                md.validFrom
            )

        if md.validTo:
            ws["{}{}".format(col.get("validTo").letter, idx)] = utils.hlpr_datetimes(
                md.validTo
            )

        if md.updateFrequency:
            ws[
                "{}{}".format(col.get("updateFrequency").letter, idx)
            ] = self.fmt.frequency_as_explicit_str(md.updateFrequency)
        if md.validityComment:
            ws[
                "{}{}".format(col.get("validityComment").letter, idx)
            ] = md.validityComment

        # -- EVENTS --------------------------------------------------------------------
        # data creation date
        if md.created:
            ws["{}{}".format(col.get("created").letter, idx)] = utils.hlpr_datetimes(
                md.created
            )

        # events count
        if md.events:
            ws["{}{}".format(col.get("events").letter, idx)] = len(md.events)

        # data last update
        if md.modified:
            ws["{}{}".format(col.get("modified").letter, idx)] = utils.hlpr_datetimes(
                md.modified
            )

        # -- TECHNICAL -----------------------------------------------------------------
        # format
        if md.format and md.type in ("rasterDataset", "vectorDataset"):
            format_lbl = next(v for k, v in md.tags.items() if "format:" in k)
            ws["{}{}".format(col.get("format").letter, idx)] = "{0} ({1} - {2})".format(
                format_lbl, md.formatVersion, md.encoding
            )
            self.stats.li_data_formats.append(format_lbl)
        elif md.format:
            ws["{}{}".format(col.get("format").letter, idx)] = "{0} {1}".format(
                md.format, md.formatVersion
            )
            self.stats.li_data_formats.append(md.format)
        else:
            pass

        # SRS
        if isinstance(md.coordinateSystem, dict):
            ws[
                "{}{}".format(col.get("coordinateSystem").letter, idx)
            ] = "{0} ({1})".format(
                md.coordinateSystem.get("name"), md.coordinateSystem.get("code")
            )

        # bounding box (envelope)
        if md.type != "resource" and md.envelope and md.envelope.get("bbox"):
            coords = md.envelope.get("coordinates")
            if md.envelope.get("type") == "Polygon":
                bbox = ",\n".join(
                    format(coord, ".4f") for coord in md.envelope.get("bbox")
                )
            elif md.envelope.get("type") == "Point":
                bbox = "Centroïde : {}{}".format(coords[0], coords[1])
            else:
                bbox = ",\n".join(
                    format(coord, ".4f") for coord in md.envelope.get("bbox")
                )
            ws["{}{}".format(col.get("envelope").letter, idx)] = bbox

        # geometry
        if md.geometry:
            ws["{}{}".format(col.get("geometry").letter, idx)] = md.geometry

        # resolution
        if md.distance:
            ws["{}{}".format(col.get("distance").letter, idx)] = md.distance

        # scale
        if md.scale:
            ws["{}{}".format(col.get("scale").letter, idx)] = md.scale

        # features objects
        if md.features:
            ws["{}{}".format(col.get("features").letter, idx)] = md.features

        # -- QUALITY -------------------------------------------------------------------
        if md.specifications:
            ws["{}{}".format(col.get("specifications").letter, idx)] = " ;\n".join(
                self.fmt.specifications(md.specifications)
            )

        # topology
        if md.topologicalConsistency:
            ws["AC{}".format(idx)] = md.topologicalConsistency

        # -- FEATURE ATTRIBUTES --------------------------------------------------------
        if md.type == "vectorDataset" and isinstance(md.featureAttributes, list):
            fields = md.featureAttributes

            # count
            ws["{}{}".format(col.get("featureAttributesCount").letter, idx)] = len(
                fields
            )
            # alphabetic list
            fields_cct = sorted(
                [
                    "{} ({}) - Type : {} - Descripion : {:.20} [...]".format(
                        field.get("name"),
                        field.get("alias"),
                        field.get("dataType"),
                        # field.get("language"),
                        field.get("description", ""),
                    )
                    for field in fields
                ]
            )
            ws["{}{}".format(col.get("featureAttributes").letter, idx)] = " ;\n".join(
                fields_cct
            )
            # if attributes analisis is activated, append fields dict
            if hasattr(self, "ws_fa"):
                self.fa_all.append(fields)
            else:
                pass

        # -- CGUs ----------------------------------------------------------------------
        if md.conditions:
            ws["{}{}".format(col.get("conditions").letter, idx)] = " ;\n".join(
                self.fmt.conditions(md.conditions)
            )

        # -- LIMITATIONS ---------------------------------------------------------------
        if md.limitations:
            ws["{}{}".format(col.get("limitations").letter, idx)] = " ;\n".join(
                self.fmt.limitations(md.limitations)
            )

        # -- CONTACTS ------------------------------------------------------------------
        if md.contacts:
            contacts = [
                "{0} ({1})".format(
                    contact.get("contact").get("name"),
                    contact.get("contact").get("email"),
                )
                for contact in md.contacts
            ]
            ws["{}{}".format(col.get("contacts").letter, idx)] = " ;\n".join(contacts)

        # -- ACTIONS -------------------------------------------------------------------
        ws["{}{}".format(col.get("hasLinkDownload").letter, idx)] = (
            "action:download" in md.tags
        )
        ws["{}{}".format(col.get("hasLinkView").letter, idx)] = "action:view" in md.tags
        ws["{}{}".format(col.get("hasLinkOther").letter, idx)] = (
            "action:other" in md.tags
        )

        # -- METADATA ------------------------------------------------------------------
        # id
        ws["{}{}".format(col.get("_id").letter, idx)] = md._id

        # creation
        if md._created:
            ws["{}{}".format(col.get("_created").letter, idx)] = utils.hlpr_datetimes(
                md._created
            )
            # add creation date (not datetime) for later stats
            self.stats.li_dates_md_created.append(
                utils.hlpr_datetimes(md._created).date()
            )

        # last update
        if md._modified:
            ws["{}{}".format(col.get("_modified").letter, idx)] = utils.hlpr_datetimes(
                md._modified
            )
            # add modification date (not datetime) for later stats, only if different from the creation date
            if md._modified != md._created:
                self.stats.li_dates_md_modified.append(
                    utils.hlpr_datetimes(md._modified).date()
                )

        # edit
        ws["{}{}".format(col.get("linkEdit").letter, idx)] = utils.get_edit_url(md)
        if self.share is not None:
            link_visu = utils.get_view_url(
                md_id=md._id, share_id=self.share._id, share_token=self.share.urlToken
            )
            ws["{}{}".format(col.get("linkView").letter, idx)] = link_visu

        # lang
        ws["{}{}".format(col.get("language").letter, idx)] = md.language

        # log
        logger.info(
            "Metadata stored: {} ({})".format(md.title_or_name(slugged=1), md._id)
        )

    # ------------ Analisis --------------------------------------------------
    def launch_analisis(self):
        """Launches special analisis, using the stats submodule."""
        # special analisis
        if hasattr(self, "ws_fa"):
            self.stats.attributes(all_attributes=self.fa_all, ws_attributes=self.ws_fa)
            # style it
            self.column_width(self.ws_fa, self.columns_attributes)
            self.row_height(self.ws_fa)
            self.styling_cells(self.ws_fa, self.columns_attributes)
            logger.info("Analisis of feature attributes sheet has been added.")

        if hasattr(self, "ws_d"):
            self.stats.pie_types(self.ws_d)
            self.stats.pie_formats(self.ws_d)
            self.stats.line_dates(self.ws_d)
            logger.info("Dashboard sheet has been added.")

    # ------------ Customize worksheet ----------------------------------------
    def headers_writer(self, ws: Worksheet, columns: ColumnPattern):
        """Writes the headers from a columns ref table to a worksheet.

        :param Worksheet ws: worksheet into write headers
        :param ColumnPattern columns: column table
        """
        # text
        for k, v in columns.items():
            if v.letter is None:
                continue
            ws["{}1".format(v.letter)] = self.tr.get(k, "Missing translation")

        # styling
        for row_cols in ws.iter_cols(min_row=1, max_row=1):
            for cell in row_cols:
                cell.style = "Headline 2"

    def column_width(self, ws: Worksheet, columns: ColumnPattern):
        """Set the width of the columns of the passed worksheet.

        :param Worksheet ws: worksheet into write headers
        :param ColumnPattern columns: column table
        """
        # apply heigth - see #52
        for v in columns.values():
            # ignore empties columns
            if v.letter is None or v.width is None:
                continue
            ws.column_dimensions[v.letter].width = v.width

    def row_height(self, ws: Worksheet, from_row: int = 2, height: int = 35):
        """Set the height of the rows of the passed worksheet.

        :param Worksheet ws: worksheet into write headers
        :param int from_row: row to start from. Default to '2' = ignoring headers.
        :param int height: fixed height to apply. Default to 35.
        """
        # apply heigth - see #52
        for i in range(from_row, ws.max_row + 1):
            ws.row_dimensions[i].height = height

    def styling_cells(self, ws: Worksheet, columns: ColumnPattern):
        """Applies the referenced style to the cells of a column.

        :param Worksheet ws: worksheet into write headers
        :param ColumnPattern columns: column table
        """
        # wrap
        for v in columns.values():
            # ignore empties columns
            if v.letter is None or v.style is None:
                continue
            # apply wrap style depending on value
            col_idx = column_index_from_string(v.letter)
            for row_cols in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                for cell in row_cols:
                    # print(cell.style)
                    cell.style = v.style

    def tunning_worksheets(self, excluded_sheets: tuple = ("dashboard",)):
        """Applies some adjustments to the sheets of the workbook: filters, frozen panels,
        print settings, etc.

        :param tuple excluded_sheets: sheets name to be excluded from the tunning.
        """
        for sheet in self.worksheets:
            # exclude specified sheets
            if sheet.title in [self.tr.get(i, "") for i in excluded_sheets]:
                logger.debug(
                    "'{}' sheet has been exclude from the tunning.".format(sheet.title)
                )
                continue
            # Freezing panes
            c_freezed = sheet["B2"]
            sheet.freeze_panes = c_freezed

            # Print properties
            sheet.print_options.horizontalCentered = True
            sheet.print_options.verticalCentered = True
            sheet.page_setup.fitToWidth = 1
            sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE

            # Others properties
            wsprops = sheet.sheet_properties
            wsprops.filterMode = True

            # enable filters
            sheet.auto_filter.ref = str("A1:{}{}").format(
                get_column_letter(sheet.max_column), sheet.max_row
            )


# #############################################################################
# ##### Stand alone program ########
# ##################################
if __name__ == "__main__":
    """ Standalone execution and development tests """
    # ------------ Specific imports ----------------
    from dotenv import load_dotenv
    from logging.handlers import RotatingFileHandler
    from os import environ
    import urllib3

    from isogeo_pysdk import Isogeo

    # ------------ Log & debug ----------------
    logger = logging.getLogger()
    logging.captureWarnings(True)
    logger.setLevel(logging.DEBUG)
    # logger.setLevel(logging.INFO)

    log_format = logging.Formatter(
        "%(asctime)s || %(levelname)s "
        "|| %(module)s - %(lineno)d ||"
        " %(funcName)s || %(message)s"
    )

    # debug to the file
    log_file_handler = RotatingFileHandler("dev_debug.log", "a", 3000000, 1)
    log_file_handler.setLevel(logging.DEBUG)
    log_file_handler.setFormatter(log_format)

    # info to the console
    log_console_handler = logging.StreamHandler()
    log_console_handler.setLevel(logging.DEBUG)
    log_console_handler.setFormatter(log_format)

    logger.addHandler(log_file_handler)
    logger.addHandler(log_console_handler)

    # ------------ Real start ----------------
    # get user ID as environment variables
    load_dotenv("dev.env")

    # ignore warnings related to the QA self-signed cert
    if environ.get("ISOGEO_PLATFORM").lower() == "qa":
        urllib3.disable_warnings()

    # for oAuth2 Backend (Client Credentials Grant) Flow
    isogeo = Isogeo(
        auth_mode="group",
        client_id=environ.get("ISOGEO_API_GROUP_CLIENT_ID"),
        client_secret=environ.get("ISOGEO_API_GROUP_CLIENT_SECRET"),
        auto_refresh_url="{}/oauth/token".format(environ.get("ISOGEO_ID_URL")),
        platform=environ.get("ISOGEO_PLATFORM", "qa"),
    )

    # getting a token
    isogeo.connect()

    # misc
    print("App used: {}".format(isogeo.app_properties.name))
    METADATA_TEST_FIXTURE_UUID = environ.get("ISOGEO_FIXTURES_METADATA_COMPLETE")
    WORKGROUP_TEST_FIXTURE_UUID = environ.get("ISOGEO_WORKGROUP_TEST_UUID")

    search = isogeo.search(
        whole_results=1,
        # query="owner:{}".format(WORKGROUP_TEST_FIXTURE_UUID),
        include="all",
    )

    isogeo.close()  # close session

    print("{}/{} metadata to be exported.".format(len(search.results), search.total))

    # instanciate th final workbook
    out_workbook = Isogeo2xlsx(
        lang=isogeo.lang, url_base_edit=isogeo.app_url, url_base_view=isogeo.oc_url,
    )
    # add needed worksheets
    out_workbook.set_worksheets(auto=search.tags.keys(), attributes=1, dashboard=1)

    # parse search results
    for md in map(Metadata.clean_attributes, search.results):
        out_workbook.store_metadatas(md)

    # launch analisis
    out_workbook.launch_analisis()

    # apply filters
    out_workbook.tunning_worksheets()

    # save file
    out_workbook.save("test_isogeo_export_to_xlsx.xlsx")

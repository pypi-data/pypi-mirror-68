# -*- coding: UTF-8 -*-
#!/usr/bin/env python

"""
    Statistics Calculator.
    Perform statistics analisis.

    Author: Isogeo

    Python: 3.7.x
    Created: 14/08/2014
"""

# ###########################################################################
# ########## Libraries ##########
# ###############################

# Standard library
from collections import Counter, defaultdict, namedtuple
from datetime import date
from itertools import zip_longest
import logging

# submodule
from isogeotoxlsx.i18n import I18N_EN, I18N_FR

# 3rd party library
from openpyxl.chart.axis import DateAxis
from openpyxl.chart import LineChart, PieChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.worksheet.worksheet import Worksheet


# ##############################################################################
# ############ Globals ############
# #################################

# LOG
logger = logging.getLogger("isogeotoxlsx")

# ############################################################################
# ######## Classes ###############
# ################################


class Stats(object):
    """Perform statistics calculations and make Excel charts."""

    li_data_formats = []
    li_dates_md_created = []
    li_dates_md_modified = []
    md_empty_fields = defaultdict(list)
    md_types_repartition = defaultdict(int)
    md_tags_occurences = defaultdict(int)

    def __init__(self, lang="fr"):
        """Instanciate stats class."""
        # self._ = _
        super(Stats, self).__init__()

        # LOCALE
        if lang.lower() == "fr":
            self.dates_fmt = "DD/MM/YYYY"
            self.locale_fmt = "fr_FR"
            self.tr = I18N_FR
        else:
            self.dates_fmt = "YYYY/MM/DD"
            self.locale_fmt = "uk_UK"
            self.tr = I18N_EN

    def attributes(self, ws_attributes: Worksheet, all_attributes: list):
        """Perform feature attributes analisis and write results into the
        wanted worksheet.

        :param Worksheet ws_attributes: sheet of a Workbook to write analisis
        :param list all_attributes: list of all feature attributes. It's a list of dicts.
        """
        idx_fa = 1
        # local arrays
        fa_names = []
        # fa_types = []
        # fa_alias = []
        # fa_descr = []

        # parsing
        for dico_fa in all_attributes:
            for fa in dico_fa:
                fa_names.append(fa.get("name"))
                # fa_alias.append(fa.get("alias", "NR"))
                # fa_types.append(fa.get("dataType"))
                # fa_descr.append(fa.get("description", "NR"))
                del fa

        # stats
        frq_names = Counter(fa_names)
        # frq_alias = Counter(fa_alias)
        # frq_types = Counter(fa_types)
        # frq_descr = Counter(fa_descr)

        # write
        ws = ws_attributes
        for fa in frq_names:
            idx_fa += 1
            ws["A{}".format(idx_fa)] = fa
            ws["B{}".format(idx_fa)] = frq_names.get(fa)

    def line_dates(
        self,
        ws: Worksheet,
        li_dates_md_created: list = None,
        li_dates_md_modified: list = None,
        cell_start_table: str = "O1",
        cell_start_chart: str = "S1",
    ):
        """Calculates metadata creation and modification dates repartition and add a \
            Line chart to the wanted sheet of Workbook.

        :param Worksheet ws: sheet of a Workbook to write analisis
        :param list li_dates_md_created: list of metadatas'creation dates. If not specified, the class attribute will be used instead.
        :param list li_dates_md_modified: list of metadatas'modification dates. If not specified, the class attribute will be used instead.
        :param str cell_start_table: cell of the sheet where to start writing table
        :param str cell_start_chart: cell of the sheet where to start writing the chart
        """
        # use passed lists or class ones
        if li_dates_md_created is None:
            li_dates_md_created = self.li_dates_md_created
        if li_dates_md_modified is None:
            li_dates_md_modified = self.li_dates_md_modified

        # compare lists
        # length
        if len(li_dates_md_created) != len(li_dates_md_modified):
            logger.warning(
                "Dates lists should have the same length. Creation: {} | Modification: {}".format(
                    len(li_dates_md_created), len(li_dates_md_modified)
                )
            )
        total_dates = len(set(li_dates_md_created + li_dates_md_modified))

        # common
        logger.debug(
            "{}/{} dates with both metadata creation and modification.".format(
                len(set(li_dates_md_created).intersection(li_dates_md_modified)),
                total_dates,
            )
        )

        # difference
        logger.debug(
            "{}/{} dates with only metadata creation.".format(
                len(set(li_dates_md_created).difference(li_dates_md_modified)),
                total_dates,
            )
        )

        logger.debug(
            "{}/{} dates with only metadata modification.".format(
                len(set(li_dates_md_modified).difference(li_dates_md_created)),
                total_dates,
            )
        )

        # use a named tuple
        DateFrequency = namedtuple(
            "DateFrequency", ["date", "count_md_created", "count_md_modified"]
        )

        # parse dates
        count_creation = Counter(li_dates_md_created)
        count_update = Counter(li_dates_md_modified)
        itr_dates_frequency = []
        for crea, mod in zip_longest(
            sorted(count_creation), sorted(count_update), fillvalue=0
        ):
            if crea == mod:
                # means a day with both metadata creation and modification
                itr_dates_frequency.append(
                    DateFrequency(crea, count_creation.get(crea), count_update.get(mod))
                )
            elif crea == 0:
                print("creation empty: {}".format(count_creation.get(crea)))
                itr_dates_frequency.append(DateFrequency(mod, 0, count_update.get(mod)))
            elif mod == 0:
                print("modification empty: {}".format(count_update.get(mod)))
                itr_dates_frequency.append(
                    DateFrequency(crea, count_creation.get(crea), 0)
                )
            else:
                itr_dates_frequency.append(
                    DateFrequency(crea, count_creation.get(crea), 0)
                )
                itr_dates_frequency.append(DateFrequency(mod, 0, count_update.get(mod)))

        # get starting cells
        min_cell_start_table = ws[cell_start_table]

        # write headers
        ws.cell(
            row=min_cell_start_table.row,
            column=min_cell_start_table.column,
            value=self.tr.get("date", "Date"),
        )
        ws.cell(
            row=min_cell_start_table.row,
            column=min_cell_start_table.column + 1,
            value="{} - {}".format(self.tr.get("occurrences"), self.tr.get("_created")),
        )
        ws.cell(
            row=min_cell_start_table.row,
            column=min_cell_start_table.column + 2,
            value="{} - {}".format(
                self.tr.get("occurrences"), self.tr.get("_modified")
            ),
        )

        # write data into worksheet
        row = min_cell_start_table.row
        for date_freq in sorted(itr_dates_frequency):
            row += 1
            ws.cell(row=row, column=min_cell_start_table.column, value=date_freq.date)
            ws.cell(
                row=row,
                column=min_cell_start_table.column + 1,
                value=date_freq.count_md_created,
            )
            ws.cell(
                row=row,
                column=min_cell_start_table.column + 2,
                value=date_freq.count_md_modified,
            )

        # Chart with date axis
        dates_chart = LineChart()

        labels = Reference(
            worksheet=ws,
            min_col=min_cell_start_table.column,
            min_row=min_cell_start_table.row + 1,
            max_row=row,
        )
        data = Reference(
            worksheet=ws,
            min_col=min_cell_start_table.column + 1,
            max_col=min_cell_start_table.column + 2,
            min_row=min_cell_start_table.row,
            max_row=row,
        )

        dates_chart.add_data(data, titles_from_data=1)
        dates_chart.set_categories(labels)

        # custom chart
        dates_chart.title = self.tr.get("date", "Date")
        # dates_chart.style = 2
        # dates_chart.smooth = True
        dates_chart.height = 10  # default is 7.5
        dates_chart.width = 30  # default is 15
        dates_chart.y_axis.title = self.tr.get("occurrences")
        dates_chart.y_axis.crossAx = 500
        dates_chart.x_axis = DateAxis(crossAx=100)
        dates_chart.x_axis.number_format = "mmm-y"
        dates_chart.x_axis.majorTimeUnit = "days"
        dates_chart.x_axis.title = "Date"

        # insert chart into the worksheet at the specified anchor
        ws.add_chart(dates_chart, cell_start_chart)

    def pie_formats(
        self,
        ws: Worksheet,
        li_formats: list = None,
        cell_start_table: str = "A20",
        cell_start_chart: str = "D20",
    ):
        """Calculates metadata formats repartition and add a Pie chart to the wanted sheet of Workbook.

        :param Worksheet ws: sheet of a Workbook to write analisis
        :param list li_formats: list of all formats labels. If not specified, the class attribute will be used instaed
        :param str cell_start_table: cell of the sheet where to start writing table
        :param str cell_start_chart: cell of the sheet where to start writing the chart
        """
        if li_formats is None:
            li_formats = self.li_data_formats

        # build the data for pie chart
        data = Counter(li_formats)

        # get starting cells
        min_cell_start_table = ws[cell_start_table]

        # write headers
        ws.cell(
            row=min_cell_start_table.row,
            column=min_cell_start_table.column,
            value=self.tr.get("format"),
        )
        ws.cell(
            row=min_cell_start_table.row,
            column=min_cell_start_table.column + 1,
            value=self.tr.get("occurrences"),
        )

        # write data into worksheet
        row = min_cell_start_table.row
        for frmt, count in data.items():
            row += 1
            ws.cell(row=row, column=min_cell_start_table.column, value=frmt.title())
            ws.cell(row=row, column=min_cell_start_table.column + 1, value=count)

        # Pie chart
        pie = PieChart()
        labels = Reference(
            worksheet=ws,
            min_col=min_cell_start_table.column,
            min_row=min_cell_start_table.row + 1,
            max_row=row,
        )
        data = Reference(
            worksheet=ws,
            min_col=min_cell_start_table.column + 1,
            min_row=min_cell_start_table.row + 1,
            max_row=row,
        )
        pie.add_data(data)
        pie.set_categories(labels)
        pie.title = self.tr.get("format") + "s"

        # Cut the first slice out of the pie
        slice = DataPoint(idx=0, explosion=20)
        pie.series[0].data_points = [slice]

        ws.add_chart(pie, cell_start_chart)

    def pie_types(
        self,
        ws: Worksheet,
        types_counters: dict = None,
        cell_start_table: str = "A1",
        cell_start_chart: str = "D1",
    ):
        """Calculates metadata types repartition and add a Pie chart to the wanted sheet of Workbook.

        :param Worksheet ws: sheet of a Workbook to write analisis
        :param dict types_counters: dictionary of types/count. If not specified, the class attribute will be used instaed
        :param str cell_start_table: cell of the sheet where to start writing table
        :param str cell_start_chart: cell of the sheet where to start writing the chart
        """
        if types_counters is None:
            types_counters = self.md_types_repartition

        # get starting cells
        min_cell_start_table = ws[cell_start_table]

        # write headers
        ws.cell(
            row=min_cell_start_table.row,
            column=min_cell_start_table.column,
            value=self.tr.get("type"),
        )
        ws.cell(
            row=min_cell_start_table.row,
            column=min_cell_start_table.column + 1,
            value=self.tr.get("occurrences"),
        )

        # write data into worksheet
        row = min_cell_start_table.row
        for md_type, count in self.md_types_repartition.items():
            row += 1
            ws.cell(
                row=row, column=min_cell_start_table.column, value=self.tr.get(md_type)
            )
            ws.cell(row=row, column=min_cell_start_table.column + 1, value=count)

        # Pie chart
        pie = PieChart()
        labels = Reference(
            worksheet=ws,
            min_col=min_cell_start_table.column,
            min_row=min_cell_start_table.row + 1,
            max_row=row,
        )
        data = Reference(
            worksheet=ws,
            min_col=min_cell_start_table.column + 1,
            min_row=min_cell_start_table.row + 1,
            max_row=row,
        )
        pie.add_data(data)
        pie.set_categories(labels)
        pie.title = self.tr.get("type") + "s"

        # Cut the first slice out of the pie
        slice = DataPoint(idx=0, explosion=20)
        pie.series[0].data_points = [slice]

        ws.add_chart(pie, cell_start_chart)


# ############################################################################
# ###### Stand alone program ########
# ###################################
if __name__ == "__main__":
    """Standalone execution and tests."""
    from openpyxl import Workbook

    # logs for debug
    logging.basicConfig(level=logging.DEBUG)

    # this module
    app = Stats()
    # workbook
    wb = Workbook()

    # types of metadatas
    ws_types = wb.create_sheet(title="Types")
    app.md_types_repartition = {
        "raster": 50,
        "resource": 10,
        "service": 40,
        "vector": 100,
    }

    app.pie_types(ws_types)

    # formats of source datasets
    ws_formats = wb.create_sheet(title="Formats")
    app.li_data_formats = [
        "PostGIS",
        "WFS",
        "PostGIS",
        "WMS",
        "Esri Shapefiles",
        "Esri Shapefiles",
        "Esri Shapefiles",
        "Esri Shapefiles",
        "Esri Shapefiles",
    ]

    app.pie_formats(
        ws_formats,
        # cell_start_table="A"  # you can specify where to write table
    )

    # creation and modification dates
    ws_history = wb.create_sheet(title="History")
    app.li_dates_md_created = [
        date(2019, 1, 1),
        date(2019, 2, 1),
        date(2019, 1, 12),
        date(2019, 1, 12),
        # date(2019, 1, 12),
        date(2019, 2, 14),
        date(2019, 2, 14),
        # date(2019, 2, 14),
        # date(2019, 2, 14),
        date(2019, 2, 28),
        # date(2019, 3, 1),
        # date(2019, 3, 2),
        # date(2019, 3, 3),
        # date(2019, 3, 4),
        # date(2019, 3, 5),
        # date(2019, 3, 5),
    ]

    app.li_dates_md_modified = [
        date(2019, 1, 1),
        date(2019, 2, 1),
        date(2019, 1, 12),
        # date(2019, 3, 12),
        # date(2019, 3, 12),
        # date(2019, 3, 12),
        # date(2019, 3, 12),
        # date(2019, 3, 12),
        # date(2019, 4, 12),
        # date(2019, 2, 14),
        # date(2019, 2, 28),
        # date(2019, 3, 1),
        # date(2019, 3, 2),
        date(2019, 3, 3),
        # date(2019, 2, 4),
        # date(2019, 2, 5),
        # None
    ]

    # print(app.li_dates_md_created.sort() == app.li_dates_md_modified.sort())
    # print(list(set(app.li_dates_md_created).intersection(app.li_dates_md_modified)))

    app.line_dates(ws=ws_history, cell_start_table="A1", cell_start_chart="E1")

    # write xlsx
    wb.save("test_stats_charts.xlsx")
